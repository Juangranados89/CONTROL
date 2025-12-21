import json
import re
from openpyxl import load_workbook


def parse_km(cell_value):
    if cell_value is None:
        return None
    text = str(cell_value).replace('\n', ' ').upper().strip()
    m = re.search(r'(\d[\d\.,]*)\s*KM', text)
    if not m:
        return None
    cleaned = m.group(1).replace('.', '').replace(',', '')
    try:
        return int(cleaned)
    except Exception:
        return None


def infer_unit(name, quantity_raw):
    upper = str(name or '').upper()
    # Reglas simples: fluidos se manejan en GAL (según requerimiento del usuario)
    if any(word in upper for word in ['ACEITE', 'REFRIGERANTE', 'LIQUIDO DE FRENOS', 'LÍQUIDO DE FRENOS']):
        return 'GAL'

    # Si viene decimal (1.75 / 0.75) también asumir GAL
    q = str(quantity_raw or '').strip()
    if q and ('.' in q or ',' in q):
        return 'GAL'

    return 'UND'


def extract_data(file_path, brand):
    print(f"Extracting from {file_path} ({brand})...")
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Header row is 19 (1-indexed) => 18 (0-indexed)
    header_row_idx = 19

    # Identify interval columns (H=8 onward); first 7 columns are metadata
    intervals = {}
    max_col = ws.max_column
    for col_idx in range(8, max_col + 1):
        km = parse_km(ws.cell(row=header_row_idx, column=col_idx).value)
        if km:
            intervals[col_idx] = km

    print(f"Found intervals: {list(intervals.values())}")

    routines = {km: {"items": [], "supplies": []} for km in intervals.values()}

    def cell_text(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None:
            return ''
        return str(v).strip()

    # Iterate data rows
    for r in range(header_row_idx + 1, ws.max_row + 1):
        col_a = cell_text(r, 1)
        if not col_a:
            continue

        col_b = cell_text(r, 2)  # referencia / tipo de lubricante
        qty = cell_text(r, 7) or '1'  # cantidad

        # Skip section headers (e.g. FILTROS)
        if col_a.upper() in {'FILTROS', 'TIPO DE FILTROS', 'TIPO DE ACEITE O REFRIGERANTE'}:
            continue

        # Check which intervals have 'X'
        for col_idx, km in intervals.items():
            marker = cell_text(r, col_idx).upper()
            if marker != 'X':
                continue

            # If there's something in col_b, treat as supply (matches current spreadsheet structure)
            if col_b:
                routines[km]["supplies"].append({
                    "name": col_a,
                    "reference": col_b,
                    "unit": infer_unit(col_a, qty),
                    "quantity": qty
                })
            else:
                routines[km]["items"].append({
                    "description": col_a,
                    "type": "Cambio" if "CAMBIAR" in col_a.upper() else "Inspección"
                })

    return routines


all_data = {
    "RAM": extract_data('CAMIONETAS RAM.xlsx', 'RAM'),
    "JMC": extract_data('CAMIONETAS JMC.xlsx', 'JMC')
}

with open('extracted_routines_v2.json', 'w') as f:
    json.dump(all_data, f, indent=2, ensure_ascii=False)

print("Extraction complete. Saved to extracted_routines_v2.json")
